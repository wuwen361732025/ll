from openpyxl import load_workbook

wb=load_workbook("1.xlsx",data_only=True)
#获取工作簿的名称
# print(wb.sheetnames)
wb1=wb["biao1"]
#读方式1
print(wb1["A3"].value)